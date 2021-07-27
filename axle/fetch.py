import csv
import datetime
import json
import logging
import os
import re

from openpyxl import load_workbook
from .helpers import (
    get_cached_path,
    get_config,
    get_format_dict,
    get_tracked_sheets,
    a1_to_rowcol,
    set_logging,
    update_formats,
    update_notes,
    validate_axle_project,
)


def fetch(verbose=False):
    """Update cached copies of sheets based on the XLSX spreadsheet. Do not update local copies."""
    set_logging(verbose)
    axle_dir = validate_axle_project()
    config = get_config(axle_dir)
    wb = load_workbook(config["Spreadsheet Path"])
    tracked_sheets = get_tracked_sheets(axle_dir)

    # TODO: handle renames, data validation

    # Collect the formatting and note data for each sheet while iterating over cells
    sheet_formats = {}
    sheet_notes = {}

    # Format ID to format for cell formatting
    id_to_format = get_format_dict(axle_dir)
    if id_to_format:
        # Format to format ID
        format_to_id = {json.dumps(v, sort_keys=True): k for k, v in id_to_format.items()}
        # Next ID for new formats
        format_ids = list(id_to_format.keys())
        format_ids.sort()
        next_fmt_id = int(format_ids[-1]) + 1
    else:
        format_to_id = {}
        next_fmt_id = 1

    new_sheets = []
    sheet_frozen = {}
    for sheet_title in wb.get_sheet_names():
        if sheet_title not in tracked_sheets:
            new_sheets.append(sheet_title)
        sheet = wb.get_sheet_by_name(sheet_title)
        frozen = sheet.freeze_panes
        if frozen:
            row, col = a1_to_rowcol(frozen)
            sheet_frozen[sheet_title] = {
                "row": row - 1,
                "col": col - 1,
            }
        else:
            sheet_frozen[sheet_title] = {"row": 0, "col": 0}

        rows = []
        cell_to_note = {}
        cell_to_format_id = {}
        for row in sheet.iter_rows():
            cells = []
            for cell in row:
                cells.append(cell.value)

                # Handle notes
                # These are called comments in openpyxl, but they're actually notes in Excel
                # Excel comments are not supported
                note = cell.comment
                if note:
                    cell_to_note[cell.coordinate] = {"text": note.text, "author": note.author}

                # Handle formatting
                fmt = get_cell_format(cell)
                if not fmt:
                    continue
                fmt_key = json.dumps(fmt, sort_keys=True)
                if fmt_key in format_to_id:
                    # Format already exists, assign that ID
                    fmt_id = format_to_id[fmt_key]
                else:
                    # Assign a new ID and add to master dict
                    fmt_id = next_fmt_id
                    format_to_id[fmt_key] = fmt_id
                    id_to_format[fmt_id] = fmt
                    next_fmt_id += 1
                # openpyxl doesn't accept ranges, so don't worry about ranges of formats
                # each cell gets its own entry in format.tsv
                cell_to_format_id[cell.coordinate] = fmt_id

            # Add all cells to the rows for this sheet
            rows.append(cells)

        # If the sheet had any formats or notes, add them to the master dicts
        if cell_to_format_id:
            sheet_formats[sheet_title] = cell_to_format_id
        if cell_to_note:
            sheet_notes[sheet_title] = cell_to_note

        # Write to cached copy
        cached_path = get_cached_path(axle_dir, sheet_title)
        with open(cached_path, "w") as fw:
            writer = csv.writer(fw, delimiter="\t", lineterminator="\n")
            writer.writerows(rows)

    # Get updated sheet details
    sheet_rows = []
    for st, details in tracked_sheets.items():
        details["Title"] = st
        details["Frozen Rows"] = sheet_frozen[st]["row"]
        details["Frozen Columns"] = sheet_frozen[st]["col"]
        sheet_rows.append(details)
    for ns in new_sheets:
        filepath = get_new_path(ns, config["Directory"], config["File Format"])
        logging.info(f"Adding new sheet '{ns}' with local path {filepath}")
        sheet_rows.append(
            {
                "Title": ns,
                "Path": filepath,
                "Frozen Rows": sheet_frozen[ns]["row"],
                "Frozen Columns": sheet_frozen[ns]["col"],
            }
        )

    # Rewrite formats JSON with new dict
    with open(f"{axle_dir}/formats.json", "w") as f:
        f.write(json.dumps(id_to_format, sort_keys=True, indent=4))
    # Update config files for formats and notes
    update_formats(axle_dir, sheet_formats)
    update_notes(axle_dir, sheet_notes)

    # Update sheet.tsv
    with open(f"{axle_dir}/sheet.tsv", "w") as fw:
        writer = csv.DictWriter(
            fw,
            delimiter="\t",
            lineterminator="\n",
            fieldnames=["Title", "Path", "Frozen Rows", "Frozen Columns"],
        )
        writer.writeheader()
        writer.writerows(sheet_rows)


def get_attributes(o):
    """Get the attributes of an object. Some of these attributes may return objects,
    so we recurse until we have the full dict."""
    fmt = {}
    for attr in dir(o):
        if attr.startswith("_"):
            continue
        try:
            v = getattr(o, attr)
        except NotImplementedError:
            # e.g., "tagname" does not work here for some objects
            continue
        if not v or callable(v):
            continue
        if attr in ["color", "fgColor", "bgColor"]:
            # Color is an object, but we just need the RGB value & tint OR a theme number
            if isinstance(v.rgb, str):
                fmt[attr] = {"rgb": v.rgb, "tint": v.tint}
            if isinstance(v.theme, int):
                fmt[attr] = {"theme": v.theme}
        elif (
            isinstance(v, str) or isinstance(v, float) or isinstance(v, int) or isinstance(v, bool)
        ):
            # Primitive datatype - just add this value
            fmt[attr] = v
        else:
            # Object - get the attributes as dict
            more_attrs = get_attributes(v)
            if more_attrs:
                fmt[attr] = more_attrs
    return fmt


def get_cell_format(cell):
    """Return the cell format as a dictionary, or None if the cell does not have a style."""
    if not cell.has_style:
        return None
    fmt = {}
    for attr in ["alignment", "border", "fill", "font", "hyperlink", "number_format"]:
        v = getattr(cell, attr)
        if not v or callable(v):
            continue
        if attr == "hyperlink":
            if isinstance(v.target, str):
                fmt[attr] = v.target
        elif (
            isinstance(v, str) or isinstance(v, float) or isinstance(v, int) or isinstance(v, bool)
        ):
            fmt[attr] = v
        else:
            more_attrs = get_attributes(v)
            if attr == "fill" and not more_attrs.get("patternType"):
                # Special case: no fill, but openpyxl returns #000 as fill color with null pattern
                fmt[attr] = {}
            else:
                fmt[attr] = more_attrs
    return fmt


def get_new_path(sheet_title, directory, file_format):
    """Create a distinct local sheet path for a sheet."""
    basename = re.sub(r"[^A-Za-z0-9]+", "_", sheet_title.lower()).strip("_")
    if directory:
        basename = os.path.join(directory, basename)
    # Make sure the path is unique - the user can change this later
    if os.path.exists(basename + ".tsv"):
        # Append datetime if this path already exists
        td = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"{basename}_{td}.{file_format}"
    else:
        filepath = basename + "." + file_format
    return filepath
