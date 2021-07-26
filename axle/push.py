import csv
import logging
import os

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from .helpers import (
    col_to_a1,
    get_cached_path,
    get_config,
    get_format_dict,
    get_sheet_formats,
    get_sheet_notes,
    get_tracked_sheets,
    set_logging,
    validate_axle_project,
)


def apply_format(cell, fmt):
    """Apply a format to a cell. The format dictionary must be turned into specific cell styles."""
    alignment = fmt.get("alignment")
    if alignment:
        alignment_obj = Alignment(**alignment)
        cell.alignment = alignment_obj

    border = fmt.get("border")
    if border:
        border_copy = border.copy()
        for side in ["left", "right", "top", "bottom", "diagonal", "vertical", "horizontal"]:
            s = border.get(side, {})
            if not s:
                continue
            s_copy = s.copy()
            if "color" in s:
                s_copy["color"] = Color(**s["color"])
            side_obj = Side(**s_copy)
            border_copy[side] = side_obj
        border_obj = Border(**border_copy)
        cell.border = border_obj

    fill = fmt.get("fill")
    if fill:
        bg_color = fill.get("bgColor", {})
        fg_color = fill.get("fgColor", {})
        fill_obj = PatternFill(
            patternType=fill.get("patternType", "solid"),
            bgColor=Color(**bg_color),
            fgColor=Color(**fg_color),
        )
        cell.fill = fill_obj

    font = fmt.get("font")
    if font:
        font_color = font.get("color", {})
        font_copy = font.copy()
        font_copy["color"] = Color(**font_color)
        font_obj = Font(**font_copy)
        cell.font = font_obj

    hyperlink = fmt.get("hyperlink")
    if hyperlink:
        cell.hyperlink = hyperlink

    number_format = fmt.get("number_format")
    if number_format:
        cell.number_format = number_format


def clear_xlsx_sheets(wb, tracked_sheets):
    """Clear all data from XLSX sheets and return a map of sheet title -> sheet obj."""
    xlsx_sheets = {}
    for sheet_title in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheet_title)
        idx = wb.sheetnames.index(sheet_title)
        wb.remove_sheet(sheet)
        if sheet_title in tracked_sheets:
            # Only recreate if we're tracking this sheet
            sheet = wb.create_sheet(sheet_title, idx)
            xlsx_sheets[sheet_title] = sheet
    return xlsx_sheets


def push_data(axle_dir, wb, tracked_sheets):
    """Push all tracked sheets to the spreadsheet. Update sheets in AXLE tracked directory. Return
    updated rows for sheet.tsv."""
    sheet_rows = []
    sheet_formats = get_sheet_formats(axle_dir)
    sheet_notes = get_sheet_notes(axle_dir)
    id_to_format = get_format_dict(axle_dir)
    for sheet_title, details in tracked_sheets.items():
        sheet_path = details["Path"]
        delimiter = "\t"
        if sheet_path.endswith(".csv"):
            delimiter = ","
        rows = []
        cols = 0
        if not os.path.exists(sheet_path):
            logging.warning(f"'{sheet_title}' exists in XLSX but has not been pulled")
            continue

        with open(sheet_path, "r") as fr:
            reader = csv.reader(fr, delimiter=delimiter)
            cached_sheet = get_cached_path(axle_dir, sheet_title)
            with open(cached_sheet, "w") as fw:
                writer = csv.writer(fw, delimiter="\t", lineterminator="\n")
                for row in reader:
                    writer.writerow(row)
                    row_len = len(row)
                    if row_len > cols:
                        cols = row_len
                    rows.append(row)

        logging.info(f"pushing data from {sheet_path} to XLSX sheet '{sheet_title}'")
        sheet = wb.create_sheet(sheet_title)
        details["Title"] = sheet_title
        sheet_rows.append(details)

        cell_formats = sheet_formats.get(sheet_title, {})
        cell_notes = sheet_notes.get(sheet_title, {})

        for row in range(0, len(rows)):
            for col in range(0, cols):
                value = rows[row][col]
                cell = sheet.cell(column=col + 1, row=row + 1, value=value)
                fmt_id = cell_formats.get(cell.coordinate)
                if fmt_id:
                    fmt = id_to_format.get(fmt_id)
                    if not fmt:
                        logging.error("Unknown format ID: " + str(fmt_id))
                    apply_format(cell, fmt)
                note = cell_notes.get(cell.coordinate)
                if note:
                    cell.comment = Comment(note["text"], note["author"])

        # Add frozen rows & cols
        frozen_row = int(details["Frozen Rows"]) + 1
        frozen_col = col_to_a1(int(details["Frozen Columns"]) + 1)
        sheet.freeze_panes = frozen_col + str(frozen_row)
    return sheet_rows


def push(verbose=False):
    """Push TSV/CSV tables to XLSX spreadsheet as sheets. Only the sheets in sheet.tsv will be
    pushed. If a sheet in the spreadsheet does not exist in sheet.tsv, it will be removed. Any sheet
    in sheet.tsv that does not exist in the spreadsheet will be created."""
    set_logging(verbose)
    axle_dir = validate_axle_project()
    config = get_config(axle_dir)

    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name("Sheet"))

    tracked_sheets = get_tracked_sheets(axle_dir)
    sheet_rows = push_data(axle_dir, wb, tracked_sheets)
    wb.save(config["Spreadsheet Path"])

    # TODO: push formats, notes, validation
    with open(f"{axle_dir}/sheet.tsv", "w") as f:
        writer = csv.DictWriter(
            f,
            delimiter="\t",
            lineterminator="\n",
            fieldnames=["Title", "Path", "Frozen Rows", "Frozen Columns",],
        )
        writer.writeheader()
        writer.writerows(sheet_rows)
